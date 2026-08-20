from __future__ import annotations

import csv
import hashlib
import io
import re
import tempfile
import uuid
from pathlib import Path
from typing import Any, Protocol

from ..common import utc_now
from .schedule_service import DEFAULT_PERIODS, ScheduleService


WEEKDAYS = {
    "周一": 1, "星期一": 1, "一": 1,
    "周二": 2, "星期二": 2, "二": 2,
    "周三": 3, "星期三": 3, "三": 3,
    "周四": 4, "星期四": 4, "四": 4,
    "周五": 5, "星期五": 5, "五": 5,
    "周六": 6, "星期六": 6, "六": 6,
    "周日": 7, "星期日": 7, "星期天": 7, "日": 7,
}
HEADER_ALIASES = {
    "name": {"课程", "课程名", "课程名称", "name"},
    "teacher": {"教师", "老师", "任课教师", "teacher"},
    "location": {"地点", "教室", "上课地点", "location"},
    "weekday": {"星期", "周几", "weekday", "day"},
    "periods": {"节次", "上课节次", "period", "periods"},
    "weeks": {"周次", "上课周次", "weeks"},
    "start_time": {"开始时间", "上课时间", "start_time"},
    "end_time": {"结束时间", "下课时间", "end_time"},
}


class OCRUnavailableError(RuntimeError):
    pass


class OCREngine(Protocol):
    def recognize(self, path: Path) -> list[dict[str, Any]]: ...


class PaddleOCREngine:
    def __init__(self) -> None:
        try:
            from paddleocr import PaddleOCR
        except ImportError as exc:
            raise OCRUnavailableError(
                "图片识别组件尚未安装。请在 planner 环境执行 pip install -r requirements-ocr.txt"
            ) from exc
        self.engine = PaddleOCR(
            lang="ch",
            use_doc_orientation_classify=False,
            use_doc_unwarping=False,
            use_textline_orientation=False,
        )

    def recognize(self, path: Path) -> list[dict[str, Any]]:
        blocks: list[dict[str, Any]] = []
        if hasattr(self.engine, "predict"):
            results = self.engine.predict(str(path))
            for result in results:
                data = getattr(result, "json", result)
                if callable(data):
                    data = data()
                if isinstance(data, dict) and "res" in data:
                    data = data["res"]
                texts = data.get("rec_texts", []) if isinstance(data, dict) else []
                scores = data.get("rec_scores", []) if isinstance(data, dict) else []
                boxes = data.get("rec_boxes", []) if isinstance(data, dict) else []
                for index, text in enumerate(texts):
                    blocks.append({
                        "text": str(text),
                        "confidence": float(scores[index]) if index < len(scores) else 0.5,
                        "bbox": list(boxes[index]) if index < len(boxes) else [],
                    })
        else:
            results = self.engine.ocr(str(path))
            for page in results or []:
                for line in page or []:
                    box, value = line
                    blocks.append({"text": str(value[0]), "confidence": float(value[1]), "bbox": box})
        return blocks


def _header_map(headers: list[Any]) -> dict[str, int]:
    mapped: dict[str, int] = {}
    for index, value in enumerate(headers):
        normalized = str(value or "").strip().lower()
        for field, aliases in HEADER_ALIASES.items():
            if normalized in aliases:
                mapped[field] = index
    return mapped


def _period_range(value: Any) -> tuple[int, int]:
    numbers = [int(item) for item in re.findall(r"\d+", str(value or ""))]
    if not numbers:
        raise ValueError("无法识别节次")
    return numbers[0], numbers[-1]


def _week_rule(value: Any) -> tuple[int, int, str, list[int]]:
    text = str(value or "1-18周").replace(" ", "")
    numbers = [int(item) for item in re.findall(r"\d+", text)]
    start, end = (numbers[0], numbers[-1]) if numbers else (1, 18)
    if "单" in text:
        return start, end, "odd", []
    if "双" in text:
        return start, end, "even", []
    if any(mark in text for mark in (",", "，", "、")) and len(numbers) > 1:
        return min(numbers), max(numbers), "custom", numbers
    return start, end, "all", []


def _weekday(value: Any) -> int:
    text = str(value or "").strip()
    if text.isdigit() and 1 <= int(text) <= 7:
        return int(text)
    for label, number in WEEKDAYS.items():
        if label in text:
            return number
    raise ValueError("无法识别星期")


class ScheduleImportService:
    def __init__(self, db_path: Path | str, ocr_engine: OCREngine | None = None) -> None:
        self.schedule = ScheduleService(db_path)
        self.repository = self.schedule.repository
        self.ocr_engine = ocr_engine

    @staticmethod
    def validate_file(filename: str, content: bytes) -> str:
        extension = Path(filename).suffix.lower()
        if extension not in {".png", ".jpg", ".jpeg", ".xlsx", ".csv"}:
            raise ValueError("仅支持 PNG、JPG、XLSX 和 CSV 课表")
        if not content:
            raise ValueError("课表文件为空")
        if len(content) > 10 * 1024 * 1024:
            raise ValueError("课表文件不能超过 10 MB")
        signatures = {
            ".png": content.startswith(b"\x89PNG\r\n\x1a\n"),
            ".jpg": content.startswith(b"\xff\xd8\xff"),
            ".jpeg": content.startswith(b"\xff\xd8\xff"),
            ".xlsx": content.startswith(b"PK"),
        }
        if extension in signatures and not signatures[extension]:
            raise ValueError("文件内容与扩展名不一致")
        return extension[1:]

    def preview(self, filename: str, content: bytes, config: dict[str, Any]) -> dict[str, Any]:
        source_type = self.validate_file(filename, content)
        source_hash = hashlib.sha256(content).hexdigest()
        semester_values = dict(config.get("semester") or {})
        semester_values.setdefault("periods", DEFAULT_PERIODS)
        semester = self.schedule.normalize_semester(semester_values)
        raw = self._parse(source_type, content)
        courses: list[dict[str, Any]] = []
        warnings: list[str] = []
        for index, item in enumerate(raw):
            try:
                draft = self._normalize_draft(item, semester, index)
            except ValueError as exc:
                draft = {
                    "draft_id": str(uuid.uuid4()),
                    "selected": False,
                    "name": str(item.get("name") or f"未识别课程 {index + 1}"),
                    "teacher": str(item.get("teacher") or ""),
                    "location": str(item.get("location") or ""),
                    "meetings": [],
                    "confidence": float(item.get("confidence", 0)),
                    "errors": [str(exc)],
                    "warnings": [],
                }
            if draft["confidence"] < 0.75:
                draft["warnings"].append("识别置信度较低，请人工核对")
            courses.append(draft)
        if self.repository.source_exists(source_type, source_hash):
            warnings.append("该源文件已经导入过，所有条目已默认取消选择")
            for course in courses:
                course["selected"] = False
                course["warnings"].append("可能是重复导入")
        self._mark_conflicts(courses, warnings, semester["id"])
        return {
            "source_type": source_type,
            "source_hash": source_hash,
            "semester": semester,
            "courses": courses,
            "warnings": warnings,
        }

    def _parse(self, source_type: str, content: bytes) -> list[dict[str, Any]]:
        if source_type == "csv":
            text = content.decode("utf-8-sig")
            return self._parse_rows(list(csv.reader(io.StringIO(text))))
        if source_type == "xlsx":
            try:
                from openpyxl import load_workbook
            except ImportError as exc:
                raise ValueError("XLSX 支持未安装，请重新安装 requirements.txt") from exc
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            sheet = workbook.active
            return self._parse_rows([list(row) for row in sheet.iter_rows(values_only=True)])
        suffix = ".png" if source_type == "png" else ".jpg"
        with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as temporary:
            temporary.write(content)
            path = Path(temporary.name)
        try:
            engine = self.ocr_engine or PaddleOCREngine()
            blocks = engine.recognize(path)
            return self._parse_ocr_blocks(blocks)
        finally:
            path.unlink(missing_ok=True)

    def _parse_rows(self, rows: list[list[Any]]) -> list[dict[str, Any]]:
        rows = [row for row in rows if any(value not in (None, "") for value in row)]
        if not rows:
            return []
        mapping = _header_map(rows[0])
        if {"name", "weekday", "periods"}.issubset(mapping):
            output = []
            for row in rows[1:]:
                def value(field: str, default: Any = "") -> Any:
                    index = mapping.get(field)
                    return row[index] if index is not None and index < len(row) else default
                output.append({
                    "name": value("name"), "teacher": value("teacher"),
                    "location": value("location"), "weekday": value("weekday"),
                    "periods": value("periods"), "weeks": value("weeks", "1-18周"),
                    "start_time": value("start_time"), "end_time": value("end_time"),
                    "confidence": 1.0,
                })
            return output
        weekday_columns = {index: _weekday(value) for index, value in enumerate(rows[0]) if any(label in str(value or "") for label in WEEKDAYS)}
        if not weekday_columns:
            raise ValueError("未找到课表表头，请使用包含课程、星期、节次列的表格")
        output = []
        for row in rows[1:]:
            period_value = row[0] if row else ""
            for index, weekday in weekday_columns.items():
                if index >= len(row) or not str(row[index] or "").strip():
                    continue
                text = str(row[index]).strip()
                lines = [part.strip() for part in re.split(r"[\n;；]", text) if part.strip()]
                output.append({
                    "name": lines[0], "teacher": lines[1] if len(lines) > 1 else "",
                    "location": lines[2] if len(lines) > 2 else "",
                    "weekday": weekday, "periods": period_value,
                    "weeks": next((line for line in lines if "周" in line), "1-18周"),
                    "confidence": 0.9,
                })
        return output

    def _parse_ocr_blocks(self, blocks: list[dict[str, Any]]) -> list[dict[str, Any]]:
        output = []
        for block in blocks:
            text = str(block.get("text") or "").strip()
            if not text or not any(label in text for label in WEEKDAYS):
                continue
            parts = [part.strip() for part in re.split(r"[|｜;；\n]", text) if part.strip()]
            try:
                weekday = _weekday(text)
                periods = next(part for part in parts if "节" in part)
            except (ValueError, StopIteration):
                continue
            weeks = next((part for part in parts if "周" in part and "周一" not in part), "1-18周")
            ignored = {periods, weeks}
            candidates = [part for part in parts if part not in ignored and not any(label in part for label in WEEKDAYS)]
            output.append({
                "name": candidates[0] if candidates else "",
                "teacher": candidates[1] if len(candidates) > 1 else "",
                "location": candidates[2] if len(candidates) > 2 else "",
                "weekday": weekday, "periods": periods, "weeks": weeks,
                "confidence": float(block.get("confidence", 0.5)),
            })
        if not output:
            def center(block: dict[str, Any]) -> tuple[float, float]:
                box = block.get("bbox") or []
                if len(box) == 4 and all(isinstance(item, (int, float)) for item in box):
                    return (float(box[0] + box[2]) / 2, float(box[1] + box[3]) / 2)
                if box and isinstance(box[0], (list, tuple)):
                    return (
                        sum(float(point[0]) for point in box) / len(box),
                        sum(float(point[1]) for point in box) / len(box),
                    )
                return (0, 0)
            headers = []
            periods = []
            for block in blocks:
                text = str(block.get("text") or "").strip()
                try:
                    if text in WEEKDAYS or text.startswith(("周", "星期")):
                        headers.append((*center(block), _weekday(text), block))
                        continue
                except ValueError:
                    pass
                if "节" in text and re.search(r"\d", text):
                    try:
                        periods.append((*center(block), _period_range(text), block))
                    except ValueError:
                        pass
            if headers and periods:
                groups: dict[tuple[int, int, int], list[dict[str, Any]]] = {}
                ignored = {id(item[3]) for item in headers} | {id(item[3]) for item in periods}
                for block in blocks:
                    if id(block) in ignored or not str(block.get("text") or "").strip():
                        continue
                    x, y = center(block)
                    header = min(headers, key=lambda item: abs(item[0] - x))
                    period = min(periods, key=lambda item: abs(item[1] - y))
                    groups.setdefault((header[2], period[2][0], period[2][1]), []).append(block)
                for (weekday, start_period, end_period), group in groups.items():
                    group.sort(key=lambda item: center(item)[1])
                    texts = [str(item.get("text") or "").strip() for item in group]
                    weeks = next((text for text in texts if "周" in text), "1-18周")
                    details = [text for text in texts if text != weeks]
                    if not details:
                        continue
                    output.append({
                        "name": details[0], "teacher": details[1] if len(details) > 1 else "",
                        "location": details[2] if len(details) > 2 else "",
                        "weekday": weekday, "periods": f"{start_period}-{end_period}节",
                        "weeks": weeks,
                        "confidence": min(float(item.get("confidence", 0.5)) for item in group),
                    })
        if not output:
            raise ValueError("未能从图片中识别课程，请裁剪无关区域或改用 XLSX/CSV")
        return output

    def _normalize_draft(self, item: dict[str, Any], semester: dict[str, Any], index: int) -> dict[str, Any]:
        name = str(item.get("name") or "").strip()
        if not name:
            raise ValueError("课程名称不能为空")
        start_period, end_period = _period_range(item.get("periods"))
        start_week, end_week, pattern, custom = _week_rule(item.get("weeks"))
        periods = {entry["period"]: entry for entry in semester["periods"]}
        start_time = str(item.get("start_time") or periods.get(start_period, {}).get("start_time") or "")
        end_time = str(item.get("end_time") or periods.get(end_period, {}).get("end_time") or "")
        meeting = self.schedule.normalize_meeting({
            "weekday": _weekday(item.get("weekday")), "start_period": start_period,
            "end_period": end_period, "start_time": start_time, "end_time": end_time,
            "start_week": start_week, "end_week": end_week,
            "week_pattern": pattern, "custom_weeks": custom,
        }, course_id="preview")
        meeting.pop("course_id")
        return {
            "draft_id": str(uuid.uuid4()), "selected": True, "name": name,
            "teacher": str(item.get("teacher") or "").strip(),
            "location": str(item.get("location") or "").strip(),
            "color": "#4f77bb", "notes": "", "meetings": [meeting],
            "confidence": float(item.get("confidence", 1)), "errors": [], "warnings": [],
        }

    def _mark_conflicts(self, drafts: list[dict[str, Any]], warnings: list[str], semester_id: str) -> None:
        existing = self.repository.meetings_between()
        for draft in drafts:
            for meeting in draft.get("meetings", []):
                duplicates = [old for old in existing if
                    old["semester_id"] == semester_id
                    and old["name"] == draft["name"]
                    and old["weekday"] == meeting["weekday"]
                    and old["start_time"] == meeting["start_time"]
                    and old["end_time"] == meeting["end_time"]
                    and old["start_week"] == meeting["start_week"]
                    and old["end_week"] == meeting["end_week"]
                    and old["week_pattern"] == meeting["week_pattern"]]
                if duplicates:
                    draft["selected"] = False
                    draft["warnings"].append("与已有课程规则完全相同，已取消选择")
                    continue
                overlap = any(
                    old["semester_id"] == semester_id
                    and old["weekday"] == meeting["weekday"]
                    and old["start_time"] < meeting["end_time"]
                    and meeting["start_time"] < old["end_time"]
                    for old in existing
                )
                if overlap:
                    draft["warnings"].append("与已有课程时间重叠")
                    warnings.append(f"{draft['name']} 与已有课程时间重叠")

    def confirm(self, payload: dict[str, Any]) -> list[str]:
        source_type = str(payload.get("source_type") or "")
        source_hash = str(payload.get("source_hash") or "")
        if source_type not in {"png", "jpg", "jpeg", "xlsx", "csv"} or len(source_hash) != 64:
            raise ValueError("导入来源无效，请重新生成预览")
        if self.repository.source_exists(source_type, source_hash):
            raise ValueError("该课表文件已经导入")
        semester = self.schedule.normalize_semester(payload.get("semester") or {})
        selected = [item for item in payload.get("courses", []) if item.get("selected", True)]
        if not selected:
            raise ValueError("至少选择一门有效课程")
        import_id = str(uuid.uuid4())
        now = utc_now()
        courses = []
        for item in selected:
            course_id = str(uuid.uuid4())
            name = str(item.get("name") or "").strip()
            if not name:
                raise ValueError("课程名称不能为空")
            meetings_source = item.get("meetings")
            if not isinstance(meetings_source, list) or not meetings_source:
                raise ValueError(f"课程 {name} 缺少有效上课时段")
            meetings = [self.schedule.normalize_meeting(meeting, course_id=course_id) for meeting in meetings_source]
            course = {
                "id": course_id, "semester_id": semester["id"], "import_id": import_id,
                "name": name, "teacher": str(item.get("teacher") or "").strip(),
                "location": str(item.get("location") or "").strip(),
                "color": str(item.get("color") or "#4f77bb"),
                "notes": str(item.get("notes") or "").strip(),
                "created_at": now, "updated_at": now, "deleted_at": None,
            }
            courses.append((course, meetings))
        return self.repository.create_import(
            {"id": import_id, "source_type": source_type, "source_hash": source_hash, "imported_at": now},
            semester,
            courses,
        )
